import numpy
import copy
import fps
import mediapipe as mp
import cv2
import openpyxl
import random


mpHands = mp.solutions.hands
mpHandModel1 = mpHands.Hands(
    model_complexity=1,
    max_num_hands=1,
    min_detection_confidence=0.90,
    min_tracking_confidence=0.05,
)
mpHandModel2 = mpHands.Hands(
    model_complexity=1,
    max_num_hands=1,
    min_detection_confidence=0.90,
    min_tracking_confidence=0.05,
)
mpHandModel3 = mpHands.Hands(
    model_complexity=1,
    max_num_hands=1,
    min_detection_confidence=0.90,
    min_tracking_confidence=0.05,
)
mpDraw = mp.solutions.drawing_utils


FPS = fps.fps()
MAX_FPS = 60
# cap = cv2.VideoCapture(0)
cap = cv2.VideoCapture("http://192.168.10.100:4747/video")
cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)


def mpLandmarks2ndarray(landmarks, correctScale=True, imgWidth=None, imgHigh=None):
    result1 = numpy.zeros((21, 3))

    if correctScale and not (imgHigh and imgWidth):
        raise Exception("input is not complete")
        return None

    try:
        for i in range(21):
            result1[i][0] = landmarks[i].x
            result1[i][1] = landmarks[i].y
            result1[i][2] = landmarks[i].z

        if correctScale:
            imgSize = (imgWidth + imgHigh) / 2
            for i in range(21):
                result1[i][0] = landmarks[i].x * imgWidth / imgSize
                result1[i][1] = landmarks[i].y * imgHigh / imgSize
                result1[i][2] = landmarks[i].z * imgWidth / imgSize

    except:
        pass

    return result1


def addNoise(img, p=5):
    result = numpy.int16(img)
    randomNum = numpy.random.randint(-p, p + 1, result.shape, "int16")
    # print("noise adding")
    result += randomNum
    result = numpy.fmax(0, result)
    result = numpy.fmin(255, result)
    result = numpy.uint8(result)
    # print("noise added")
    return result


img2fixed = False

wb = openpyxl.Workbook()
wb.create_sheet("test")
s1 = wb["test"]
# s1.cell(1, 1).value = "normal"
# s1.cell(1, 3).value = "fixed"
s1.cell(1, 5).value = "fixed + noised"
row = 2

while True:
    curFps = FPS.get(limitFps=MAX_FPS)
    avgFps = FPS.avgFps()
    # print(round(avgFps))

    handLandmark1 = None
    handLandmark2 = None
    handLandmark3 = None

    ret, img = cap.read()

    # print("min",imgDiff.max())
    # print(sum(sum(imgDiff))/(imgDiff.shape[0]*imgDiff.shape[1]))

    if ret:
        imgHigh = img.shape[0]
        imgWidth = img.shape[1]
        imgSize = (imgHigh + imgWidth) / 2

        imgRGB = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        # img1 = copy.deepcopy(imgRGB)
        if not img2fixed:
            img2 = copy.deepcopy(imgRGB)
        # img3 = copy.deepcopy(imgRGB)
        img3 = addNoise(img2)

        # img1show = cv2.cvtColor(img1, cv2.COLOR_RGB2BGR)
        # img2show = cv2.cvtColor(img2, cv2.COLOR_RGB2BGR)
        img3show = cv2.cvtColor(img3, cv2.COLOR_RGB2BGR)

        # result1 = mpHandModel1.process(img1)
        # result2 = mpHandModel2.process(img2)
        result3 = mpHandModel3.process(img3)

        # if result1.multi_hand_landmarks:

        #     handLandmark1 = mpLandmarks2ndarray(
        #         landmarks=result1.multi_hand_landmarks[0].landmark,
        #         correctScale=True,
        #         imgWidth=imgWidth,
        #         imgHigh=imgHigh,
        #     )

        #     try:
        #         for handLms in result1.multi_hand_landmarks:
        #             mpDraw.draw_landmarks(img1show, handLms, mpHands.HAND_CONNECTIONS)
        #     except:
        #         pass

        # if result2.multi_hand_landmarks:

        #     handLandmark2 = mpLandmarks2ndarray(
        #         landmarks=result2.multi_hand_landmarks[0].landmark,
        #         correctScale=True,
        #         imgWidth=imgWidth,
        #         imgHigh=imgHigh,
        #     )

        #     try:
        #         for handLms in result2.multi_hand_landmarks:
        #             mpDraw.draw_landmarks(img2show, handLms, mpHands.HAND_CONNECTIONS)
        #     except:
        #         pass

        if result3.multi_hand_landmarks:

            handLandmark3 = mpLandmarks2ndarray(
                landmarks=result3.multi_hand_landmarks[0].landmark,
                correctScale=True,
                imgWidth=imgWidth,
                imgHigh=imgHigh,
            )

            try:
                for handLms in result3.multi_hand_landmarks:
                    mpDraw.draw_landmarks(img3show, handLms, mpHands.HAND_CONNECTIONS)
            except:
                pass

        cv2.putText(
            img3show,
            "fps: " + str(int(curFps)),
            (0, 50),
            cv2.FONT_HERSHEY_SIMPLEX,
            2,
            (0, 0, 0),
            2,
        )
        # cv2.imshow("img1", img1show)
        # cv2.imshow("img2", img2show)
        cv2.imshow("img3", img3show)
        # cv2.imshow("diff", img3show-img2show)

    cv2KeyEvent = cv2.waitKey(1)

    if cv2KeyEvent == ord("r") or img2fixed:
        # s1.cell(row, 1).value = handLandmark1[9][0]
        # s1.cell(row, 2).value = handLandmark1[9][1]
        # s1.cell(row, 3).value = handLandmark2[9][0]
        # s1.cell(row, 4).value = handLandmark2[9][1]
        s1.cell(row, 5).value = handLandmark3[9][0]
        s1.cell(row, 6).value = handLandmark3[9][1]
        row += 1
        print(row)

    if cv2KeyEvent == ord("f"):
        img2fixed = True
    if cv2KeyEvent == ord("q"):
        break
    if cv2KeyEvent == 27:
        break

wb.save("dataGen/" + "test.xlsx")
