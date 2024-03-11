import time

runTimeRecorder = time.perf_counter()
print("loading dictionary")

import cv2
import mediapipe as mp
import matplotlib
import numpy
import copy
import matplotlib.pyplot as plt
import openpyxl

import fps
import tools
import draw3DHand

MAX_FPS = 60
CAM_NUM = 0
MAX_HANDS_AMOUNT = 1

# 3d圖xyz比例記得調一樣
# 不然手會爛掉
#
#

time.perf_counter()
print("took " + str(time.perf_counter() - runTimeRecorder) + " sec")
runTimeRecorder = time.perf_counter()
print("loading camera")

cap = cv2.VideoCapture(CAM_NUM)

if cap is None or not cap.isOpened():
    raise Exception(
        'Unable to access camera, please check README.md for more info')

print("took " + str(time.perf_counter() - runTimeRecorder) + " sec")
runTimeRecorder = time.perf_counter()
print("loading model")

mpHands = mp.solutions.hands
handModel = mpHands.Hands(model_complexity=1,
                          max_num_hands=MAX_HANDS_AMOUNT,
                          min_detection_confidence=0.95,
                          min_tracking_confidence=0.05)
mpDraw = mp.solutions.drawing_utils

print("took " + str(time.perf_counter() - runTimeRecorder) + " sec")
runTimeRecorder = time.perf_counter()

hancLandmarkConnection = numpy.array([
    [0, 1],
    [1, 2],
    [2, 3],
    [3, 4],
    [0, 5],
    [5, 6],
    [6, 7],
    [7, 8],
    [9, 10],
    [10, 11],
    [11, 12],
    [13, 14],
    [14, 15],
    [15, 16],
    [0, 17],
    [17, 18],
    [18, 19],
    [19, 20],
    [5, 9],
    [9, 13],
    [13, 17],
])

hand3D = draw3DHand.handDrawer3D()

FPS = fps.fps()
startRecord = False
standardRelativeHandLandmark = numpy.zeros((21,3))



wb = openpyxl.Workbook()
wb.create_sheet("yaw")
s1 = wb['yaw']            # 開啟工作表


while (True):

    curFps = FPS.get()

    ret, img = cap.read()

    if (ret):
        imgHigh = img.shape[0]
        imgWidth = img.shape[1]
        imgSize = (imgHigh + imgWidth) / 2

        imgRGB = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        handResult = handModel.process(imgRGB)
        if (handResult.multi_hand_landmarks):
            handLandmark = numpy.zeros((21, 3))

            for i in range(21):
                handLandmark[i][0] = (handResult.multi_hand_landmarks[0].
                                      landmark[i].x) * imgWidth / imgSize
                handLandmark[i][1] = (handResult.multi_hand_landmarks[0].
                                      landmark[i].y) * imgHigh / imgSize
                handLandmark[i][2] = (handResult.multi_hand_landmarks[0].
                                      landmark[i].z) * imgWidth / imgSize

            vector_a = handLandmark[13] - handLandmark[0]
            vector_b = handLandmark[13] - handLandmark[5]
            handFaceVector = tools.externalProduct(vector_a, vector_b)
            handFaceVector = handFaceVector

            handFaceYaw = tools.getDegree(
                numpy.array([handFaceVector[0], handFaceVector[2]]),
                numpy.array([ 0, -1]))
            if (handFaceVector[0] < 0):
                handFaceYaw = -handFaceYaw

            handFacePitch = tools.getDegree(
                numpy.array([handFaceVector[0],handFaceVector[1],handFaceVector[2]]),
                numpy.array([handFaceVector[0],               0 ,handFaceVector[2]]))
            if (handFaceVector[1] < 0):
                handFacePitch = -handFacePitch

            handSize = tools.getVectorLength(handLandmark[5] -
                                             handLandmark[17])

            relativeLandmarkPos = numpy.zeros((21, 3))
            for i in range(21):
                relativeLandmarkPos[i] = (handLandmark[i] -
                                          handLandmark[13]) / handSize

                relativeLandmarkPos[i][0], relativeLandmarkPos[i][
                    2] = tools.rotateVector(
                        numpy.array([
                            relativeLandmarkPos[i][0],
                            relativeLandmarkPos[i][2]
                        ]), -handFaceYaw)

                relativeLandmarkPos[i][1], relativeLandmarkPos[i][
                    2] = tools.rotateVector(
                        numpy.array([
                            relativeLandmarkPos[i][1],
                            relativeLandmarkPos[i][2]
                        ]), -handFacePitch)

            handFaceRowVector = relativeLandmarkPos[13]-relativeLandmarkPos[0]
            handFaceRow = tools.getDegree(handFaceRowVector,numpy.array([0,-1,0]))
            if(handFaceRowVector[0]<0):
                handFaceRow = -handFaceRow


            # print(relativeLandmarkPos[13], relativeLandmarkPos[0])

            for i in range(21):
                relativeLandmarkPos[i][0], relativeLandmarkPos[i][
                    1] = tools.rotateVector(
                        numpy.array([
                            relativeLandmarkPos[i][0],
                            relativeLandmarkPos[i][1]
                        ]), -handFaceRow)



            cv2KeyEvent = cv2.waitKey(1)
            if (cv2KeyEvent == ord('s') and not standardRelativeHandLandmark.any()):
                standardRelativeHandLandmark = relativeLandmarkPos
                s1.cell(1,1).value = "standerd"
                s1.cell(1,2).value = "yaw"
                s1.cell(1,3).value = "pitch"
                s1.cell(1,4).value = "row"
                for i in range(21):
                    s1.cell(1,5+i).value = "landmark " + str(i)
                
            if (cv2KeyEvent == ord('r') and standardRelativeHandLandmark.any()):
                print(relativeLandmarkPos - standardRelativeHandLandmark)



            s1['A1'].value = 'apple'     # 儲存格 A1 內容為 apple
            s1['A2'].value = 'orange'    # 儲存格 A2 內容為 orange
            s1['A3'].value = 'banana'    # 儲存格 A3 內容為 banana
            s1.cell(1,2).value = 100     # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
            s1.cell(2,2).value = 200     # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
            s1.cell(3,2).value = 300     # 儲存格 B3 內容 ( row=3, column=2 ) 為 300













            # print(handFaceVector)
            print(handFaceYaw, handFacePitch, handFaceRow)

            # print(relativeLandmarkPos[4])
            hand3D.draw(relativeLandmarkPos)

            for handLms in handResult.multi_hand_landmarks:
                mpDraw.draw_landmarks(img, handLms, mpHands.HAND_CONNECTIONS)

        cv2.putText(img, "fps: " + str(int(curFps)), (0, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 0), 2)
        cv2.imshow("cam", img)
    else:
        pass

    cv2KeyEvent = cv2.waitKey(1)
    if (cv2KeyEvent == ord('q')):
        break

    if (cv2KeyEvent == 27):
        break

wb.save('yaw.xlsx')